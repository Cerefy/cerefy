"""πX Excel Intelligence Engine.

Understands Excel workbooks beyond simple sheet reading: workbooks, hidden sheets,
named ranges, tables, formulas, column business meaning, data quality, and
inter-sheet relationships.
"""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.worksheet.table import Table as OpenpyxlTable
from openpyxl.worksheet.worksheet import Worksheet

from app.cognitive_data_layer.excel.models import (
    ExcelColumnInsight,
    ExcelFormula,
    ExcelNamedRange,
    ExcelQualityScore,
    ExcelRelationship,
    ExcelSheet,
    ExcelTable,
    ExcelUnderstandingResult,
)
from app.cognitive_data_layer.quality import DataQualityEngine
from app.cognitive_data_layer.schema import SchemaDiscoverer
from app.cognitive_data_layer.semantic import SemanticUnderstandingEngine


class ExcelUnderstandingEngine:
    """Analyze Excel workbooks and extract business intelligence."""

    def __init__(self) -> None:
        self._schema = SchemaDiscoverer()
        self._semantic = SemanticUnderstandingEngine()
        self._quality = DataQualityEngine()

    def analyze(
        self,
        source: str | Path | bytes,
        file_name: str | None = None,
    ) -> ExcelUnderstandingResult:
        """Analyze an Excel workbook and return a structured understanding."""
        wb = self._load_workbook(source)
        if not file_name:
            file_name = str(source) if isinstance(source, (str, Path)) else "workbook.xlsx"

        named_ranges = [self._extract_named_range(name) for name in wb.defined_names.values()]
        hidden_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "hidden"]

        sheets: list[ExcelSheet] = []
        all_tables: list[ExcelTable] = []
        for index, ws in enumerate(wb.worksheets, start=1):
            sheet = self._analyze_sheet(ws, index)
            sheets.append(sheet)
            all_tables.extend(sheet.tables)

        relationships = self._discover_relationships(all_tables)
        quality = self._compute_quality(sheets)
        business_concepts = sorted({
            concept
            for sheet in sheets
            for table in sheet.tables
            for concept in (c.business_concept for c in table.columns if c.business_concept)
        })

        return ExcelUnderstandingResult(
            file_name=file_name,
            sheet_count=len(sheets),
            hidden_sheets=hidden_sheets,
            named_ranges=named_ranges,
            sheets=sheets,
            relationships=relationships,
            quality=quality,
            business_concepts=business_concepts,
        )

    # ------------------------------------------------------------------ #
    # Workbook loading
    # ------------------------------------------------------------------ #
    @staticmethod
    def _load_workbook(source: str | Path | bytes) -> openpyxl.Workbook:
        if isinstance(source, (str, Path)):
            return openpyxl.load_workbook(source, data_only=False)
        return openpyxl.load_workbook(BytesIO(source), data_only=False)

    # ------------------------------------------------------------------ #
    # Sheet analysis
    # ------------------------------------------------------------------ #
    def _analyze_sheet(self, ws: Worksheet, index: int) -> ExcelSheet:
        df = pd.DataFrame(ws.values)
        if df.empty:
            return ExcelSheet(name=ws.title, index=index, hidden=ws.sheet_state == "hidden")

        # Promote first row to headers if it looks like headers
        df = self._promote_headers(df)
        df = df.replace({np.nan: None})

        row_count, column_count = df.shape
        tables = self._extract_tables(ws, df)
        # If no explicit table objects, treat the whole sheet as one table.
        if not tables:
            tables = [self._analyze_table(ws.title, f"{ws.title}_data", df)]

        return ExcelSheet(
            name=ws.title,
            index=index,
            hidden=ws.sheet_state == "hidden",
            tables=tables,
            row_count=row_count,
            column_count=column_count,
        )

    def _promote_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        if len(df) == 0:
            return df
        first_row = df.iloc[0]
        if first_row.notna().all() and first_row.astype(str).str.strip().ne("").all():
            df = df.iloc[1:].reset_index(drop=True)
            df.columns = [str(c).strip() for c in first_row]
        else:
            df.columns = [str(c) for c in df.columns]
        return df

    def _extract_tables(self, ws: Worksheet, df: pd.DataFrame) -> list[ExcelTable]:
        tables: list[ExcelTable] = []
        for table_obj in ws.tables.values():
            if isinstance(table_obj, OpenpyxlTable):
                table_df = self._extract_table_range(df, table_obj, ws)
                tables.append(
                    self._analyze_table(ws.title, table_obj.name, table_df, is_table_object=True)
                )
        return tables

    def _extract_table_range(
        self,
        df: pd.DataFrame,
        table_obj: OpenpyxlTable,
        ws: Worksheet,
    ) -> pd.DataFrame:
        # openpyxl table ref is like "A1:D10"
        ref = table_obj.ref
        try:
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
            if min_col and min_row and max_col and max_row:
                # DataFrame is 0-indexed; openpyxl rows are 1-indexed.
                start_row = min_row - 1
                end_row = max_row - 1
                start_col = min_col - 1
                end_col = max_col
                subset = df.iloc[start_row : end_row + 1, start_col:end_col]
                return self._promote_headers(subset)
        except Exception:  # noqa: BLE001
            pass
        return df

    def _analyze_table(
        self,
        sheet_name: str,
        table_name: str,
        df: pd.DataFrame,
        is_table_object: bool = False,
    ) -> ExcelTable:
        df = df.dropna(how="all").reset_index(drop=True)
        if df.empty:
            return ExcelTable(name=table_name, sheet=sheet_name)

        columns = self._analyze_columns(df)
        formulas = self._extract_formulas(df)
        return ExcelTable(
            name=table_name,
            sheet=sheet_name,
            columns=columns,
            row_count=len(df),
            formulas=formulas,
            is_table_object=is_table_object,
        )

    def _analyze_columns(self, df: pd.DataFrame) -> list[ExcelColumnInsight]:
        columns: list[ExcelColumnInsight] = []
        for col in df.columns:
            series = df[col].dropna()
            sample_values = series.head(5).astype(str).tolist()
            mapping = self._semantic.infer_entity(str(col), sample_values)
            inferred_dtype = self._infer_dtype(series)
            business_concept = (
                mapping.entity_type.value if mapping.entity_type.value != "unknown" else None
            )

            columns.append(
                ExcelColumnInsight(
                    name=str(col),
                    data_type=inferred_dtype,
                    semantic_type=mapping.entity_type.value,
                    business_concept=business_concept,
                    sample_values=sample_values,
                    null_count=int(df[col].isna().sum()),
                    unique_count=int(series.nunique()),
                    confidence=round(mapping.confidence, 2),
                )
            )
        return columns

    @staticmethod
    def _infer_dtype(series: pd.Series) -> str:
        if pd.api.types.is_datetime64_any_dtype(series):
            return "datetime"
        if pd.api.types.is_integer_dtype(series):
            return "integer"
        if pd.api.types.is_float_dtype(series):
            return "float"
        if pd.api.types.is_bool_dtype(series):
            return "boolean"
        return "string"

    # ------------------------------------------------------------------ #
    # Formula intelligence
    # ------------------------------------------------------------------ #
    def _extract_formulas(self, df: pd.DataFrame) -> list[ExcelFormula]:
        formulas: list[ExcelFormula] = []
        # Formula cells are not available in data_only=False mode via ws.values,
        # so we scan the dataframe text for formulas starting with '='.
        for col_idx, col in enumerate(df.columns):
            for row_idx, value in df.iloc[:, col_idx].items():
                if isinstance(value, str) and value.startswith("="):
                    cell = f"{openpyxl.utils.get_column_letter(col_idx + 1)}{row_idx + 2}"
                    formulas.append(
                        ExcelFormula(
                            cell=cell,
                            formula=value,
                            functions=self._extract_functions(value),
                            dependencies=self._extract_dependencies(value),
                        )
                    )
        return formulas

    @staticmethod
    def _extract_functions(formula: str) -> list[str]:
        return sorted(set(re.findall(r"([A-Z][A-Z0-9_]*)\s*\(", formula)))

    @staticmethod
    def _extract_dependencies(formula: str) -> list[str]:
        # Capture simple cell/range references like A1, A1:B10, Sheet!A1
        pattern = r"(?:[A-Za-z_][A-Za-z0-9_]*!)?\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?"
        return sorted(set(re.findall(pattern, formula)))

    # ------------------------------------------------------------------ #
    # Named ranges
    # ------------------------------------------------------------------ #
    def _extract_named_range(self, name: Any) -> ExcelNamedRange:
        try:
            return ExcelNamedRange(
                name=str(name.name),
                value=str(name.value) if name.value else None,
                refers_to=str(name.attr_text) if name.attr_text else None,
            )
        except Exception:  # noqa: BLE001
            return ExcelNamedRange(name=str(name))

    # ------------------------------------------------------------------ #
    # Relationship discovery
    # ------------------------------------------------------------------ #
    def _discover_relationships(self, tables: list[ExcelTable]) -> list[ExcelRelationship]:
        relationships: list[ExcelRelationship] = []
        id_columns = [
            (table.name, col.name)
            for table in tables
            for col in table.columns
            if self._is_identifier_column(col.name)
        ]
        for source_table, source_col in id_columns:
            source_values = self._get_column_values(tables, source_table, source_col)
            for target_table, target_col in id_columns:
                if target_table == source_table and target_col == source_col:
                    continue
                target_values = self._get_column_values(tables, target_table, target_col)
                overlap = source_values & target_values
                if len(overlap) >= 2:
                    relationships.append(
                        ExcelRelationship(
                            source_table=source_table,
                            source_column=source_col,
                            target_table=target_table,
                            target_column=target_col,
                            relationship_type="foreign_key",
                            confidence=round(min(1.0, len(overlap) / 10), 2),
                        )
                    )
        return relationships

    @staticmethod
    def _is_identifier_column(name: str) -> bool:
        lower = name.lower()
        return "id" in lower or "code" in lower or "number" in lower

    @staticmethod
    def _get_column_values(tables: list[ExcelTable], table_name: str, column_name: str) -> set[str]:
        for table in tables:
            if table.name == table_name:
                for col in table.columns:
                    if col.name == column_name:
                        return set(col.sample_values)
        return set()

    # ------------------------------------------------------------------ #
    # Quality scoring
    # ------------------------------------------------------------------ #
    def _compute_quality(self, sheets: list[ExcelSheet]) -> ExcelQualityScore:
        total_cells = 0
        null_cells = 0
        duplicate_rows = 0
        total_rows = 0
        formula_count = 0

        for sheet in sheets:
            for table in sheet.tables:
                row_count = max(1, table.row_count)
                col_count = max(1, len(table.columns))
                total_cells += row_count * col_count
                null_cells += sum(c.null_count for c in table.columns)
                duplicate_rows += self._estimate_duplicates(table)
                total_rows += row_count
                formula_count += len(table.formulas)

        if total_cells == 0:
            return ExcelQualityScore()

        completeness = max(0.0, min(1.0, 1.0 - (null_cells / total_cells)))
        uniqueness = max(0.0, min(1.0, 1.0 - (duplicate_rows / max(1, total_rows))))
        # Consistency bonus for formulas and semantic confidence
        confidence_sum = sum(
            c.confidence
            for sheet in sheets
            for table in sheet.tables
            for c in table.columns
        )
        col_count = max(1, sum(len(t.tables) for t in sheets))
        avg_confidence = confidence_sum / col_count
        consistency = min(1.0, 0.5 + (0.5 * avg_confidence))
        if formula_count > 0:
            consistency = min(1.0, consistency + 0.05)

        overall = round(min(1.0, completeness * 0.4 + uniqueness * 0.3 + consistency * 0.3), 2)
        issue_count = null_cells + duplicate_rows
        return ExcelQualityScore(
            overall_score=overall,
            completeness=round(completeness, 2),
            uniqueness=round(uniqueness, 2),
            consistency=round(consistency, 2),
            issue_count=issue_count,
        )

    @staticmethod
    def _estimate_duplicates(table: ExcelTable) -> int:
        # Approximate duplicates by comparing sample values across columns.
        seen: set[tuple[str, ...]] = set()
        duplicates = 0
        for col in table.columns:
            sample = tuple(sorted(col.sample_values[:5]))
            if sample in seen:
                duplicates += 1
            seen.add(sample)
        return duplicates


async def analyze_excel(
    source: str | Path | bytes,
    file_name: str | None = None,
) -> ExcelUnderstandingResult:
    """Convenience async wrapper for ExcelUnderstandingEngine.analyze."""
    return ExcelUnderstandingEngine().analyze(source, file_name=file_name)
