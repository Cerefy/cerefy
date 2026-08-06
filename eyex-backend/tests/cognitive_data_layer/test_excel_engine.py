"""Tests for the πX Excel Intelligence Engine."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from app.cognitive_data_layer.excel import ExcelUnderstandingEngine


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "sales.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active worksheet")
    ws.title = "Customers"
    ws.append(["CustomerID", "CustomerName", "Revenue"])
    ws.append(["C1", "Acme Corp", 1000])
    ws.append(["C2", "Globex", 2500])

    orders = wb.create_sheet("Orders")
    orders.append(["OrderID", "CustomerID", "Amount"])
    orders.append(["O1", "C1", 500])
    orders.append(["O2", "C2", 750])
    orders.append(["O3", "C1", 500])

    hidden = wb.create_sheet("Notes")
    hidden.sheet_state = "hidden"
    hidden.append(["Note"])
    hidden.append(["Secret note"])

    # Named range
    from openpyxl.workbook.defined_name import DefinedName

    defined_name = DefinedName("CustomerRevenue", attr_text="Customers!$C$2:$C$3")
    wb.defined_names.add(defined_name)

    wb.save(path)
    return path


class TestExcelUnderstandingEngine:
    def test_analyze_workbook(self, sample_workbook: Path) -> None:
        engine = ExcelUnderstandingEngine()
        result = engine.analyze(sample_workbook)

        assert result.file_name.endswith("sales.xlsx")
        assert result.sheet_count == 3
        assert "Notes" in result.hidden_sheets
        assert any(nr.name == "CustomerRevenue" for nr in result.named_ranges)

    def test_discovers_business_concepts(self, sample_workbook: Path) -> None:
        engine = ExcelUnderstandingEngine()
        result = engine.analyze(sample_workbook)

        customers = next(s for s in result.sheets if s.name == "Customers")
        concepts = {c.name: c.business_concept for c in customers.tables[0].columns}
        assert concepts.get("CustomerID") is not None
        assert concepts.get("Revenue") is not None
        assert "revenue" in result.business_concepts

    def test_discovers_relationships(self, sample_workbook: Path) -> None:
        engine = ExcelUnderstandingEngine()
        result = engine.analyze(sample_workbook)

        rel = next(
            (r for r in result.relationships if r.source_column == "CustomerID"),
            None,
        )
        assert rel is not None
        assert rel.target_column == "CustomerID"

    def test_quality_score(self, sample_workbook: Path) -> None:
        engine = ExcelUnderstandingEngine()
        result = engine.analyze(sample_workbook)

        assert result.quality.overall_score > 0
        assert 0.0 <= result.quality.completeness <= 1.0
        assert 0.0 <= result.quality.uniqueness <= 1.0
        assert 0.0 <= result.quality.consistency <= 1.0

    def test_formula_extraction(self, tmp_path: Path) -> None:
        path = tmp_path / "formula.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("No active worksheet")
        ws.append(["A", "B", "Total"])
        ws.append([1, 2, "=SUM(A2:B2)"])
        ws.append([3, 4, "=SUM(A3:B3)"])
        wb.save(path)

        engine = ExcelUnderstandingEngine()
        result = engine.analyze(path)
        formulas = [f for sheet in result.sheets for table in sheet.tables for f in table.formulas]
        assert len(formulas) == 2
        assert "SUM" in formulas[0].functions

    def test_analyze_bytes(self, sample_workbook: Path) -> None:
        engine = ExcelUnderstandingEngine()
        with open(sample_workbook, "rb") as f:
            data = f.read()
        result = engine.analyze(data, file_name="sales.xlsx")
        assert result.file_name == "sales.xlsx"
        assert result.sheet_count == 3
